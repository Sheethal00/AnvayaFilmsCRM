#!/usr/bin/env python3
"""Assemble dist/AnvayaFilmsCRM.xlsm from schema/*.yaml + vba/*.

Two separate concerns, handled honestly rather than papered over:

1. Sheet structure (headers, column widths, data validation dropdowns
   sourced from 19_Settings lists, currency/date number formats,
   conditional formatting on Status columns) — built cross-platform with
   openpyxl. This part always runs.

2. VBA module import (vba/modules/*.bas, vba/classes/*.cls) — openpyxl
   cannot write a vbaProject stream into a workbook that doesn't already
   have one; there is no pure-Python way to compile/inject VBA. This step
   only runs if Excel is available via COM automation (Windows, pywin32
   installed), matching the CI approach noted in CLAUDE.md's testing
   philosophy. Elsewhere, it's skipped with an explicit warning rather
   than silently producing a workbook someone thinks has working macros.

Run:
    python scripts/build_workbook.py --out dist/AnvayaFilmsCRM.xlsm
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

REPO_ROOT = Path(__file__).resolve().parent.parent
SCHEMA_DIR = REPO_ROOT / "schema"
VBA_MODULES_DIR = REPO_ROOT / "vba" / "modules"
VBA_CLASSES_DIR = REPO_ROOT / "vba" / "classes"

# Real Excel formulas for schema/*.yaml columns of type "formula", written
# into row 2 of each sheet as a template. Sheets listed here are converted
# to Excel Tables (see apply_formula_row) so that when a user adds a row
# directly beneath the table in Excel, the formula autofills down —
# that's an Excel Table behavior, not something this script can trigger
# for rows added after the file is built. References use plain relative
# cell addressing (row 2), not structured [@Column] references, since
# that's what Excel itself writes when you fill a formula down a table
# and is simpler to keep in sync with the column layout comments in the
# VBA modules.
FORMULA_TEMPLATES: dict[str, dict[str, str]] = {
    "05_Bookings": {
        "BalanceAmount": (
            "=E2-SUMIFS('12_Payments'!D:D,'12_Payments'!B:B,A2,"
            '\'12_Payments\'!H:H,"Cleared")'
        ),
        "DeliveryDeadline": "=I2+DefaultDeliveryDays",
    },
    "03_Quotation_Manager": {
        "Subtotal": "=SUMIF('03a_Quotation_LineItems'!A:A,A2,'03a_Quotation_LineItems'!E:E)",
        "GST%": "=GSTRate",
        "GSTAmount": "=E2*F2",
        "Total": "=E2+G2",
    },
    "03a_Quotation_LineItems": {
        "LineTotal": "=C2*D2",
    },
    "06_Wedding_Calendar": {
        "CountdownDays": "=C2-TODAY()",
    },
    "10_Production_Pipeline": {
        "DaysInCurrentStage": "=TODAY()-C2",
    },
    "04_Client_Master": {
        "TotalLifetimeValue": "=SUMIF('05_Bookings'!B:B,A2,'05_Bookings'!E:E)",
    },
    "15_Freelancers": {
        "TotalPaidYTD": (
            "=SUMIFS('14_Expenses'!E:E,'14_Expenses'!G:G,B2,"
            '\'14_Expenses\'!C:C,"Freelancer",'
            "'14_Expenses'!B:B,\">=\"&DATE(YEAR(TODAY()),1,1))"
        ),
    },
    "16_Vendors": {
        "TotalPaidYTD": (
            "=SUMIFS('14_Expenses'!E:E,'14_Expenses'!G:G,B2,"
            "'14_Expenses'!B:B,\">=\"&DATE(YEAR(TODAY()),1,1))"
        ),
    },
}

# Excel tends to inherit "Date" display format on cells that subtract two
# dates, even when the result is a plain day-count, not a date — masking
# real values (or blank-input artifacts) as nonsensical dates. Force these
# formula columns back to a number format explicitly.
FORMULA_NUMBER_FORMATS: dict[str, dict[str, str]] = {
    "06_Wedding_Calendar": {"CountdownDays": "0"},
    "10_Production_Pipeline": {"DaysInCurrentStage": "0"},
}

# 01_Dashboard KPI tiles (SPEC.md §3 "01_Dashboard"). Not a per-row Table
# like the sheets above — a fixed label/value layout, built directly from
# this list rather than the generic per-column loop (see
# build_dashboard_sheet). Each tuple is (label, formula, number_format).
#
# "Employee Workload" deviates from SPEC.md's literal
# `Status="Active"` — 07_Team_Allocation's actual Status dropdown
# (schema/07_team_allocation.yaml) is Assigned/Confirmed/Completed/
# No-show, no "Active" value exists. Treated as Assigned + Confirmed
# (allocated but not yet Completed/No-show) since that's the closest
# real equivalent to "currently active."
DASHBOARD_KPIS: list[tuple[str, str, str | None]] = [
    ("Total Leads", "=COUNTA('02_CRM_Leads'!A:A)-1", None),
    (
        "Conversion Rate",
        '=IFERROR(COUNTIF(\'02_CRM_Leads\'!J:J,"Won")/(COUNTA(\'02_CRM_Leads\'!A:A)-1),0)',
        "0%",
    ),
    (
        "Revenue (This Month)",
        "=SUMIFS('12_Payments'!D:D,'12_Payments'!H:H,\"Cleared\","
        "'12_Payments'!E:E,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),"
        "'12_Payments'!E:E,\"<\"&EDATE(DATE(YEAR(TODAY()),MONTH(TODAY()),1),1))",
        "#,##0",
    ),
    (
        "Revenue (This Year)",
        "=SUMIFS('12_Payments'!D:D,'12_Payments'!H:H,\"Cleared\","
        "'12_Payments'!E:E,\">=\"&DATE(YEAR(TODAY()),1,1))",
        "#,##0",
    ),
    (
        "Outstanding Payments",
        '=SUMIFS(\'12_Payments\'!D:D,\'12_Payments\'!H:H,"Pending")',
        "#,##0",
    ),
    (
        "Upcoming Weddings (30 days)",
        "=COUNTIFS('06_Wedding_Calendar'!C:C,\">=\"&TODAY(),"
        "'06_Wedding_Calendar'!C:C,\"<=\"&TODAY()+30)",
        None,
    ),
    (
        "Employee Workload (active allocations)",
        '=COUNTIF(\'07_Team_Allocation\'!F:F,"Assigned")'
        '+COUNTIF(\'07_Team_Allocation\'!F:F,"Confirmed")',
        None,
    ),
    (
        "Equipment Utilization",
        "=IFERROR(COUNTIF('08_Equipment'!F:F,\"Assigned\")"
        "/(COUNTA('08_Equipment'!A:A)-1),0)",
        "0%",
    ),
]


def _month_start_formula(n: int) -> str:
    """First-of-month date, n months after Jan 1 of the current calendar
    year — self-adjusting off TODAY() rather than a hardcoded year, so
    13_Income/17_Accounting's 12-month grid always shows the current
    calendar year without needing a rebuild every January."""
    return f"EDATE(DATE(YEAR(TODAY()),1,1),{n})"


def _month_income_formula(n: int) -> str:
    start = _month_start_formula(n)
    end = _month_start_formula(n + 1)
    return (
        "=SUMIFS('12_Payments'!D:D,'12_Payments'!H:H,\"Cleared\","
        f"'12_Payments'!E:E,\">=\"&{start},'12_Payments'!E:E,\"<\"&{end})"
    )


def _month_expense_formula(n: int) -> str:
    start = _month_start_formula(n)
    end = _month_start_formula(n + 1)
    return (
        "=SUMIFS('14_Expenses'!E:E,'14_Expenses'!B:B,"
        f"\">=\"&{start},'14_Expenses'!B:B,\"<\"&{end})"
    )


STATUS_COLORS = {
    # Applied only to columns literally named "Status" — enough sheets
    # share that column name that a generic rule is worth it; anything
    # sheet-specific should live in that sheet's yaml instead.
    "Won": "C6EFCE",
    "Lost": "FFC7CE",
    "Cleared": "C6EFCE",
    "Pending": "FFEB9C",
    "Bounced": "FFC7CE",
    "Cancelled": "FFC7CE",
    "Delivered": "C6EFCE",
}


def load_schemas() -> list[dict]:
    schemas = []
    for path in sorted(SCHEMA_DIR.glob("*.yaml")):
        with path.open() as f:
            data = yaml.safe_load(f)
        if data and "sheet" in data:
            schemas.append(data)
    return schemas


def build_structure(wb: Workbook, schemas: list[dict]) -> dict[str, str]:
    settings = next((s for s in schemas if s["sheet"] == "19_Settings"), None)
    lists = settings.get("lists", {}) if settings else {}

    # Sort so numeric sheet prefixes come out in SPEC.md order (01, 02, 03a...).
    schemas_sorted = sorted(schemas, key=lambda s: s["sheet"])

    default_sheet = wb.active
    first = True
    named_ranges: dict[str, str] = {}

    for schema in schemas_sorted:
        sheet_name = schema["sheet"]
        if first:
            ws = default_sheet
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        if sheet_name == "19_Settings":
            named_ranges = build_settings_sheet(ws, schema)
            continue

        if sheet_name == "01_Dashboard":
            build_dashboard_sheet(ws, schema)
            continue

        if sheet_name == "13_Income":
            build_income_sheet(ws)
            continue

        if sheet_name == "17_Accounting":
            build_accounting_sheet(ws)
            continue

        if sheet_name == "18_Reports":
            build_reports_sheet(ws, lists)
            continue

        columns = schema.get("columns")
        if not columns:
            ws["A1"] = f"{sheet_name} — {schema.get('type', 'view')} sheet"
            ws["A1"].font = Font(italic=True)
            continue

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col["name"])
            cell.font = Font(bold=True)
            ws.column_dimensions[cell.column_letter].width = max(
                12, len(col["name"]) + 2
            )

            apply_data_validation(ws, col, col_idx, lists)
            apply_conditional_formatting(ws, col, col_idx)

        ws.freeze_panes = "A2"
        apply_formula_row(ws, columns, sheet_name)

    return named_ranges


def build_dashboard_sheet(ws, schema: dict) -> None:
    """01_Dashboard is read-only — no stored data, just formulas pulling
    live from other sheets (SPEC.md §3). Written directly from
    DASHBOARD_KPIS instead of the generic per-column loop, since there's
    no per-row Table semantics here (nothing to add rows to).
    """
    bold = Font(bold=True)
    ws["A1"] = "KPI"
    ws["B1"] = "Value"
    ws["A1"].font = bold
    ws["B1"].font = bold

    for row_idx, (label, formula, number_format) in enumerate(DASHBOARD_KPIS, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        cell = ws.cell(row=row_idx, column=2, value=formula)
        if number_format:
            cell.number_format = number_format

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = "A2"


def build_income_sheet(ws) -> None:
    """13_Income (SPEC.md §3) — filtered/grouped VIEW of 12_Payments where
    Status=Cleared, by month. Not a separately maintained sheet; nothing
    here is typed in by hand, it's all live formulas over 12_Payments.
    """
    bold = Font(bold=True)
    ws["A1"] = "Month"
    ws["B1"] = "Income (Cleared Payments)"
    ws["A1"].font = bold
    ws["B1"].font = bold

    for n in range(12):
        row = 2 + n
        cell_a = ws.cell(row=row, column=1, value=f"={_month_start_formula(n)}")
        cell_a.number_format = "mmm yyyy"
        cell_b = ws.cell(row=row, column=2, value=_month_income_formula(n))
        cell_b.number_format = "#,##0"

    total_row = 14
    ws.cell(row=total_row, column=1, value="Total").font = bold
    total_cell = ws.cell(row=total_row, column=2, value="=SUM(B2:B13)")
    total_cell.number_format = "#,##0"
    total_cell.font = bold

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A2"


def build_accounting_sheet(ws) -> None:
    """17_Accounting (SPEC.md §3) — Monthly P&L and Cash Flow, both rollup
    views over 12_Payments + 14_Expenses, no separately entered data.
    P&L and Cash Flow collapse to the same monthly net here because this
    is a cash-basis single ledger (no accrual/AR/AP) — the Cumulative
    Cash Flow column is what actually distinguishes the two, showing the
    running balance rather than just the monthly delta.
    """
    bold = Font(bold=True)
    headers = ["Month", "Income", "Expenses", "Net P&L", "Cumulative Cash Flow"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for n in range(12):
        row = 2 + n
        cell_a = ws.cell(row=row, column=1, value=f"={_month_start_formula(n)}")
        cell_a.number_format = "mmm yyyy"
        ws.cell(row=row, column=2, value=_month_income_formula(n)).number_format = "#,##0"
        ws.cell(row=row, column=3, value=_month_expense_formula(n)).number_format = "#,##0"
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").number_format = "#,##0"
        ws.cell(row=row, column=5, value=f"=SUM($D$2:D{row})").number_format = "#,##0"

    total_row = 14
    ws.cell(row=total_row, column=1, value="Total").font = bold
    for col_letter in ("B", "C", "D"):
        cell = ws.cell(
            row=total_row,
            column=ord(col_letter) - ord("A") + 1,
            value=f"=SUM({col_letter}2:{col_letter}13)",
        )
        cell.number_format = "#,##0"
        cell.font = bold
    ws.cell(row=total_row, column=5, value="=E13").number_format = "#,##0"
    ws.cell(row=total_row, column=5).font = bold

    for col_letter, width in (("A", 14), ("B", 16), ("C", 16), ("D", 16), ("E", 20)):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


# 18_Reports (SPEC.md §3) section list. Each entry is
# (section_title, header_row, category_values, category_formula_fn) where
# category_formula_fn(category_cell_ref) -> list[str] returns one formula
# per data column. Categories are the fixed enumerated lists these reports
# group by (schema/19_settings.yaml lists + 02_CRM_Leads.Status values) —
# bounded, so a static formula grid works instead of a real PivotTable
# (openpyxl can't author a live/refreshable one).
def build_reports_sheet(ws, lists: dict) -> None:
    bold = Font(bold=True)
    section_fill_font = Font(bold=True, size=12)
    row = 1

    def section_title(text: str) -> int:
        nonlocal row
        ws.cell(row=row, column=1, value=text).font = section_fill_font
        row += 1
        return row

    def header(*names: str) -> None:
        nonlocal row
        for col_idx, name in enumerate(names, start=1):
            ws.cell(row=row, column=col_idx, value=name).font = bold
        row += 1

    def data_row(values: list, number_format: str | None = None) -> int:
        nonlocal row
        this_row = row
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=this_row, column=col_idx, value=value)
            if number_format and col_idx > 1:
                cell.number_format = number_format
        row += 1
        return this_row

    def total_row(label: str, cols: list[str], first_data_row: int, last_data_row: int) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = bold
        for col_letter in cols:
            cell = ws.cell(
                row=row,
                column=ord(col_letter) - ord("A") + 1,
                value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})",
            )
            cell.font = bold
        row += 1

    def blank_line() -> None:
        nonlocal row
        row += 1

    # --- 1. Period Summary (covers the spec's Daily/Weekly/Monthly/
    # Quarterly/Annual ask; a full second monthly grid would just repeat
    # 17_Accounting, so this is period-to-date snapshots instead) --------
    section_title("Period Summary (Income vs Expenses, period-to-date)")
    header("Period", "Income", "Expenses", "Net")
    periods = [
        ("This Week", "TODAY()-6", "TODAY()"),
        ("This Month", "DATE(YEAR(TODAY()),MONTH(TODAY()),1)", "TODAY()"),
        (
            "This Quarter",
            "DATE(YEAR(TODAY()),INT((MONTH(TODAY())-1)/3)*3+1,1)",
            "TODAY()",
        ),
        ("This Year", "DATE(YEAR(TODAY()),1,1)", "TODAY()"),
    ]
    first_data_row = row
    for label, start, end in periods:
        income = (
            "=SUMIFS('12_Payments'!D:D,'12_Payments'!H:H,\"Cleared\","
            f"'12_Payments'!E:E,\">=\"&{start},'12_Payments'!E:E,\"<=\"&{end})"
        )
        expense = (
            "=SUMIFS('14_Expenses'!E:E,"
            f"'14_Expenses'!B:B,\">=\"&{start},'14_Expenses'!B:B,\"<=\"&{end})"
        )
        r = data_row([label, income, expense, None], number_format="#,##0")
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = "#,##0"
    blank_line()

    # --- 2. Bookings & Revenue by Package (merges the spec's separate
    # "Revenue Report" package dimension and "Package Report" — both group
    # by the same EventType/Package field on 05_Bookings, so one table
    # covers both) --------------------------------------------------------
    section_title("Bookings & Revenue by Package")
    header("Package", "Bookings", "Total Package Value")
    first_data_row = row
    for package in lists.get("PackagesList", []):
        data_row(
            [
                package,
                f"=COUNTIF('05_Bookings'!D:D,A{row})",
                f"=SUMIF('05_Bookings'!D:D,A{row},'05_Bookings'!E:E)",
            ],
            number_format="#,##0",
        )
    total_row("Total", ["B", "C"], first_data_row, row - 1)
    blank_line()

    # --- 3. Expense Report, by category -----------------------------------
    section_title("Expenses by Category")
    header("Category", "Total Amount")
    first_data_row = row
    for category in lists.get("ExpenseCategoryList", []):
        data_row(
            [category, f"=SUMIF('14_Expenses'!C:C,A{row},'14_Expenses'!E:E)"],
            number_format="#,##0",
        )
    total_row("Total", ["B"], first_data_row, row - 1)
    blank_line()

    # --- 4. Lead Report, by source and by status --------------------------
    section_title("Leads by Source")
    header("Source", "Lead Count")
    first_data_row = row
    for source in lists.get("LeadSourceList", []):
        data_row([source, f"=COUNTIF('02_CRM_Leads'!F:F,A{row})"])
    total_row("Total", ["B"], first_data_row, row - 1)
    blank_line()

    section_title("Leads by Status")
    header("Status", "Lead Count")
    # Mirrors 02_CRM_Leads.yaml's Status column values — not in
    # 19_Settings.lists since it's an inline `values:` list, not a
    # list_ref, so it isn't available via the `lists` dict here.
    lead_statuses = ["New", "Contacted", "Quoted", "Negotiating", "Won", "Lost"]
    first_data_row = row
    for status in lead_statuses:
        data_row([status, f"=COUNTIF('02_CRM_Leads'!J:J,A{row})"])
    total_row("Total", ["B"], first_data_row, row - 1)
    blank_line()

    # --- 5. Employee Performance -------------------------------------------
    # SPEC.md asks for "avg. days per stage" grouped by TeamMemberID, but
    # 10a_Pipeline_History (SPEC.md §3) has no editor/team-member column —
    # only BookingID/Stage/EnteredDate/ExitedDate — so there's no field to
    # group by per-editor. This reports avg. days per stage aggregated
    # across all bookings/editors instead; a per-editor breakdown would
    # need a schema change (an editor column on 10a_Pipeline_History).
    section_title("Average Days Per Pipeline Stage (all editors — see note)")
    ws.cell(
        row=row,
        column=1,
        value=(
            "10a_Pipeline_History has no per-stage editor column, so this "
            "can't be broken down by TeamMemberID as SPEC.md describes — "
            "aggregated across all bookings instead."
        ),
    ).font = Font(italic=True, size=9)
    row += 1
    header("Stage", "Avg Days (closed stage instances)")
    hist_range = "'10a_Pipeline_History'!"
    for stage in lists.get("PipelineStageList", []):
        formula = (
            f"=IFERROR(SUMPRODUCT(({hist_range}$B$2:$B$5000=A{row})*"
            f"({hist_range}$D$2:$D$5000<>\"\")*"
            f"({hist_range}$D$2:$D$5000-{hist_range}$C$2:$C$5000))/"
            f"SUMPRODUCT(({hist_range}$B$2:$B$5000=A{row})*"
            f"({hist_range}$D$2:$D$5000<>\"\")),0)"
        )
        data_row([stage, formula], number_format="0.0")
    blank_line()

    # --- 6. Client Report: lifetime value & repeat bookings ----------------
    # Fixed-capacity (50 rows) live pass-through of 04_Client_Master +
    # a repeat-booking count — IF-guarded so it reads blank, not 0/#REF,
    # past however many clients actually exist.
    section_title("Client Report (lifetime value & repeat bookings)")
    header("ClientID", "Client Name", "Bookings", "Lifetime Value")
    for src in range(2, 52):  # 04_Client_Master row 2 = its first client
        this_row = row
        guard = f"'04_Client_Master'!A{src}=\"\""
        data_row(
            [
                f"=IF({guard},\"\",'04_Client_Master'!A{src})",
                f"=IF({guard},\"\",'04_Client_Master'!B{src})",
                f"=IF({guard},\"\",COUNTIF('05_Bookings'!B:B,'04_Client_Master'!A{src}))",
                f"=IF({guard},\"\",'04_Client_Master'!I{src})",
            ],
            number_format=None,
        )
        ws.cell(row=this_row, column=4).number_format = "#,##0"

    for col_letter, width in (("A", 24), ("B", 16), ("C", 24), ("D", 20)):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def build_settings_sheet(ws, schema: dict) -> dict[str, str]:
    """19_Settings gets a bespoke layout instead of the generic per-column
    one, because IDGenerator.bas reads actual cells here at runtime — it
    searches column A for a counter key and reads the pattern/last-value
    from the next two columns. Without this, GenerateID() has nothing to
    read and every ID-generating macro fails immediately.

    Also writes GSTRate/DefaultDeliveryDays as named ranges so formulas
    on other sheets (see FORMULA_TEMPLATES) can reference them by name
    instead of a hardcoded cell address.
    """
    bold = Font(bold=True)

    # --- ID counters: A (key), B (pattern), C (last value, blank = none issued yet)
    ws["A1"] = "CounterKey"
    ws["B1"] = "Pattern"
    ws["C1"] = "LastValue"
    for addr in ("A1", "B1", "C1"):
        ws[addr].font = bold

    row = 2
    for counter in schema.get("id_counters", []):
        ws.cell(row=row, column=1, value=counter["key"])
        ws.cell(row=row, column=2, value=counter["pattern"])
        row += 1

    # --- config values: E (setting name), F (value)
    ws["E1"] = "Setting"
    ws["F1"] = "Value"
    ws["E1"].font = bold
    ws["F1"].font = bold

    gst_row = 2
    delivery_row = 3
    ws.cell(row=gst_row, column=5, value="GSTRate")
    ws.cell(row=gst_row, column=6, value=schema["values"]["GSTRate"]["value"])
    ws.cell(row=delivery_row, column=5, value="DefaultDeliveryDays")
    ws.cell(row=delivery_row, column=6, value=schema["values"]["DefaultDeliveryDays"]["value"])

    # --- team list: H onward, headers only — staff fill this in directly.
    team_columns = schema.get("team_list", {}).get("columns", [])
    for idx, col_name in enumerate(team_columns):
        cell = ws.cell(row=1, column=8 + idx, value=col_name)
        cell.font = bold

    for col_letter in ("A", "B", "C", "E", "F", "H", "I", "J", "K", "L"):
        ws.column_dimensions[col_letter].width = 18

    ws.freeze_panes = "A2"

    return {
        "GSTRate": f"'{ws.title}'!$F${gst_row}",
        "DefaultDeliveryDays": f"'{ws.title}'!$F${delivery_row}",
    }


def apply_formula_row(ws, columns: list[dict], sheet_name: str) -> None:
    """Writes real Excel formulas into row 2 for this sheet's formula-type
    columns (see FORMULA_TEMPLATES), then wraps header+row2 in an Excel
    Table. The Table is what makes Excel auto-fill the formula down when
    someone adds a new row directly beneath it — a plain range doesn't do
    that. Sheets with no formula columns are left as plain ranges.
    """
    templates = FORMULA_TEMPLATES.get(sheet_name)
    if not templates:
        return

    number_formats = FORMULA_NUMBER_FORMATS.get(sheet_name, {})
    for col_idx, col in enumerate(columns, start=1):
        formula = templates.get(col["name"])
        if formula:
            cell = ws.cell(row=2, column=col_idx, value=formula)
            if col["name"] in number_formats:
                cell.number_format = number_formats[col["name"]]

    last_col_letter = ws.cell(row=1, column=len(columns)).column_letter
    table_name = "tbl_" + re.sub(r"[^0-9A-Za-z_]", "_", sheet_name)
    table = Table(displayName=table_name, ref=f"A1:{last_col_letter}2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False
    )
    ws.add_table(table)


def apply_data_validation(ws, col: dict, col_idx: int, lists: dict) -> None:
    col_type = col.get("type")
    values = None

    if col_type == "dropdown":
        if "values" in col:
            values = col["values"]
        elif "list_ref" in col:
            values = lists.get(col["list_ref"])

    if not values:
        return

    formula = '"' + ",".join(str(v) for v in values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Value must be one of the allowed options."
    dv.errorTitle = "Invalid entry"
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    dv.add(f"{col_letter}2:{col_letter}1048576")
    ws.add_data_validation(dv)


def apply_conditional_formatting(ws, col: dict, col_idx: int) -> None:
    if col.get("name") != "Status":
        return
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    rng = f"{col_letter}2:{col_letter}1048576"
    for value, hex_color in STATUS_COLORS.items():
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{value}"'],
            fill=_solid_fill(hex_color),
        )
        ws.conditional_formatting.add(rng, rule)


def _solid_fill(hex_color: str):
    from openpyxl.styles import PatternFill

    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


XL_MACRO_ENABLED_WORKBOOK = 52  # Excel FileFormat enum: xlOpenXMLWorkbookMacroEnabled


def try_import_vba(xlsx_path: Path, xlsm_out: Path) -> bool:
    """Best-effort VBA import via Excel COM automation. Windows-only.

    Takes the *structure-only* .xlsx build_structure() wrote, imports the
    VBA modules into it via a live Excel instance, then lets Excel itself
    Save As macro-enabled — that's what actually writes a valid
    xl/vbaProject.bin part. openpyxl cannot do this step; see the module
    docstring.
    """
    try:
        import win32com.client  # type: ignore
    except ImportError:
        print(
            "NOTE: pywin32 not available — skipping VBA module import.\n"
            "      Structure-only workbook written as .xlsx; import\n"
            "      vba/modules/*.bas and vba/classes/*.cls manually in\n"
            "      Excel, or run this build on Windows with Excel +\n"
            "      pywin32 installed to get a working .xlsm.\n"
            "      NOTE: even on Windows, Excel's Trust Center setting\n"
            "      'Trust access to the VBA project object model' must be\n"
            "      enabled (File > Options > Trust Center > Trust Center\n"
            "      Settings > Macro Settings) or this import step fails."
        )
        return False

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path))
        vb_project = wb.VBProject

        for bas_file in sorted(VBA_MODULES_DIR.glob("*.bas")):
            vb_project.VBComponents.Import(str(bas_file))

        # .cls files are never imported as new components -- each one is
        # merged into an *existing* built-in code module: ThisWorkbook.cls
        # into the workbook's ThisWorkbook component, and every other
        # <SheetName>.cls into that worksheet's own code module (matched
        # by filename stem to sheet name). Worksheet code modules get
        # their VBA component name (e.g. "Sheet7") auto-assigned by Excel
        # when the sheet is created -- there's no way to "import" a file
        # as a specific worksheet's code-behind, only to write into the
        # one that already exists for that sheet.
        for cls_file in sorted(VBA_CLASSES_DIR.glob("*.cls")):
            code = strip_cls_header(cls_file.read_text())
            if cls_file.stem == "ThisWorkbook":
                component = vb_project.VBComponents("ThisWorkbook")
            else:
                sheet_name = cls_file.stem
                try:
                    ws = wb.Worksheets(sheet_name)
                except Exception:
                    print(f"WARNING: {cls_file.name} has no matching sheet '{sheet_name}' — skipped")
                    continue
                component = vb_project.VBComponents(ws.CodeName)
            component.CodeModule.AddFromString(code)

        xlsm_out.parent.mkdir(parents=True, exist_ok=True)
        wb.SaveAs(str(xlsm_out), FileFormat=XL_MACRO_ENABLED_WORKBOOK)
        wb.Close()
        print(f"VBA modules imported — wrote macro-enabled workbook to {xlsm_out}")
        return True
    finally:
        excel.Quit()


def strip_cls_header(text: str) -> str:
    """Drop the VERSION/BEGIN...END/Attribute header lines a .cls file
    starts with — those are metadata for a *new* class module; when
    merging into the existing built-in ThisWorkbook module we only want
    the executable code below them."""
    lines = text.splitlines()
    code_start = 0
    for i, line in enumerate(lines):
        if not line.startswith(("VERSION", "BEGIN", "END", "Attribute", "  ")):
            code_start = i
            break
    return "\n".join(lines[code_start:])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, default=REPO_ROOT / "dist" / "AnvayaFilmsCRM.xlsm")
    args = parser.parse_args()

    schemas = load_schemas()
    if not schemas:
        print("No schema files found in schema/ — nothing to build.", file=sys.stderr)
        return 1

    wb = Workbook()
    named_ranges = build_structure(wb, schemas)
    for name, ref in named_ranges.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

    # Always build structure into a genuine .xlsx first. openpyxl decides
    # whether to write macro-enabled content types purely from the file
    # extension it's given — saving straight to .xlsm here (with no
    # vba_archive set) writes a workbook that *claims* to be macro-enabled
    # and even links a vbaProject.bin relationship, but never writes that
    # part. Excel treats that as a corrupt file. Building as .xlsx first
    # sidesteps the mismatch entirely.
    xlsx_path = args.out.with_suffix(".xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"Wrote sheet structure to {xlsx_path} ({len(schemas)} sheets)")

    if args.out.suffix.lower() == ".xlsm":
        vba_embedded = try_import_vba(xlsx_path, args.out)
        if not vba_embedded:
            print(
                f"Requested {args.out.name}, but no VBA was embedded — "
                f"leaving {xlsx_path.name} as the buildable artifact "
                f"instead of writing a broken .xlsm."
            )
    return 0


if __name__ == "__main__":
    sys.exit(main())
