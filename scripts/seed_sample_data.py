#!/usr/bin/env python3
"""Populate a built workbook with fixture data from sample_data/*.csv, for
manual testing in Excel. Never used to seed real client data.

Run:
    python scripts/seed_sample_data.py dist/AnvayaFilmsCRM.xlsm
"""
from __future__ import annotations

import csv
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
SAMPLE_DATA_DIR = REPO_ROOT / "sample_data"
SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import build_workbook  # noqa: E402 -- needs sys.path tweak above

# sample_data/<file>.csv -> sheet name it seeds. 19_Settings.team_list is
# handled separately (seed_team_list) since it's a sub-table inside a
# bespoke-layout sheet, not a standalone table sheet with a plain header row.
CSV_TO_SHEET = {
    "leads.csv": "02_CRM_Leads",
    "client_master.csv": "04_Client_Master",
    "quotation_manager.csv": "03_Quotation_Manager",
    "quotation_lineitems.csv": "03a_Quotation_LineItems",
    "bookings.csv": "05_Bookings",
    "wedding_calendar.csv": "06_Wedding_Calendar",
    "team_allocation.csv": "07_Team_Allocation",
    "equipment.csv": "08_Equipment",
    "shoot_checklist.csv": "09_Shoot_Checklist",
    "production_pipeline.csv": "10_Production_Pipeline",
    "pipeline_history.csv": "10a_Pipeline_History",
    "album_tracker.csv": "11_Album_Tracker",
    "payments.csv": "12_Payments",
    "expenses.csv": "14_Expenses",
    "freelancers.csv": "15_Freelancers",
    "vendors.csv": "16_Vendors",
}

# Matches a bare row-2-relative cell reference like "A2" or "AB2" inside a
# FORMULA_TEMPLATES string (never a whole-column ref like "D:D" or a sheet
# name like '12_Payments', since those don't have a single/double letter
# immediately followed by "2" with nothing alphanumeric before it).
ROW2_REF_RE = re.compile(r"(?<![A-Za-z0-9])([A-Z]{1,2})2(?![0-9])")


def formula_for_row(template: str, row: int) -> str:
    """FORMULA_TEMPLATES formulas are written for row 2. When seeding more
    than one data row into a sheet that has a formula column (e.g. two
    quotes' worth of line items in 03a_Quotation_LineItems), rows past 2
    need the same formula with row-relative refs shifted — Excel's own
    "type in the row below a Table" auto-fill only happens for a human
    typing live, not for cells written by this script."""
    return ROW2_REF_RE.sub(lambda m: f"{m.group(1)}{row}", template)


def coerce(value: str):
    """CSV cells are always strings; write them back as real Excel types
    so formulas that do arithmetic/date-math on seeded rows (e.g.
    05_Bookings row 2's BalanceAmount/DeliveryDeadline formulas against
    the seeded BK-0001 row) actually compute instead of silently treating
    the cell as text."""
    if value.strip().upper() in ("TRUE", "FALSE"):
        return value.strip().upper() == "TRUE"
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        pass
    return value


def fill_formula_rows(ws, sheet_header: list[str], last_row: int) -> None:
    """For sheets with FORMULA_TEMPLATES, rows 3+ don't get their formula
    from build_workbook.py (which only writes row 2) or from Excel's live
    Table auto-fill (nothing is "typed" here) — so this script has to write
    them explicitly, then extend the Table's ref to cover the new rows."""
    templates = build_workbook.FORMULA_TEMPLATES.get(ws.title)
    if not templates or last_row < 3:
        return

    number_formats = build_workbook.FORMULA_NUMBER_FORMATS.get(ws.title, {})
    header_index = {name: idx for idx, name in enumerate(sheet_header, start=1)}

    for row_idx in range(3, last_row + 1):
        for col_name, template in templates.items():
            col_idx = header_index.get(col_name)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=formula_for_row(template, row_idx))
            if col_name in number_formats:
                cell.number_format = number_formats[col_name]

    table_name = "tbl_" + re.sub(r"[^0-9A-Za-z_]", "_", ws.title)
    if table_name in ws.tables:
        last_col_letter = ws.cell(row=1, column=len(sheet_header)).column_letter
        ws.tables[table_name].ref = f"A1:{last_col_letter}{last_row}"


def seed_sheet(ws, csv_path: Path) -> int:
    with csv_path.open(newline="") as f:
        reader = csv.reader(f)
        header = next(reader)

        sheet_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if header != sheet_header:
            raise ValueError(
                f"{csv_path.name} header {header} doesn't match "
                f"{ws.title} sheet header {sheet_header} — "
                f"update schema/*.yaml or the CSV to match."
            )

        row_count = 0
        for row_idx, row in enumerate(reader, start=2):
            for col_idx, value in enumerate(row, start=1):
                if value == "":
                    # Leave blank rather than overwrite -- row 2 on some
                    # sheets (e.g. 05_Bookings) carries a live formula in
                    # formula-type columns (see build_workbook.py
                    # FORMULA_TEMPLATES); an empty CSV cell there means
                    # "don't touch this column", not "clear it".
                    continue
                ws.cell(row=row_idx, column=col_idx, value=coerce(value))
            row_count += 1

        fill_formula_rows(ws, sheet_header, last_row=1 + row_count)
        return row_count


def seed_team_list(ws, csv_path: Path) -> int:
    """19_Settings.team_list lives at columns H:L of the 19_Settings sheet
    (see build_workbook.py build_settings_sheet) — headers only, no data
    rows, since staff normally fill it in directly in Excel. Seeds sample
    rows the same way, starting at H2."""
    with csv_path.open(newline="") as f:
        reader = csv.reader(f)
        header = next(reader)

        sheet_header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, min_col=8, max_col=12))]
        if header != sheet_header:
            raise ValueError(
                f"{csv_path.name} header {header} doesn't match "
                f"19_Settings team_list header {sheet_header} — "
                f"update schema/19_settings.yaml or the CSV to match."
            )

        row_count = 0
        for row_idx, row in enumerate(reader, start=2):
            for col_offset, value in enumerate(row):
                if value == "":
                    continue
                ws.cell(row=row_idx, column=8 + col_offset, value=coerce(value))
            row_count += 1
        return row_count


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: seed_sample_data.py <path-to-xlsm>", file=sys.stderr)
        return 1

    workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        print(f"{workbook_path} does not exist — build it first with build_workbook.py", file=sys.stderr)
        return 1

    wb = load_workbook(workbook_path, keep_vba=workbook_path.suffix == ".xlsm")

    for csv_name, sheet_name in CSV_TO_SHEET.items():
        csv_path = SAMPLE_DATA_DIR / csv_name
        if not csv_path.exists():
            print(f"skip: {csv_path} not found")
            continue
        if sheet_name not in wb.sheetnames:
            print(f"skip: sheet '{sheet_name}' not in {workbook_path}")
            continue
        n = seed_sheet(wb[sheet_name], csv_path)
        print(f"seeded {n} row(s) into {sheet_name} from {csv_name}")

    team_list_csv = SAMPLE_DATA_DIR / "team_list.csv"
    if team_list_csv.exists() and "19_Settings" in wb.sheetnames:
        n = seed_team_list(wb["19_Settings"], team_list_csv)
        print(f"seeded {n} row(s) into 19_Settings.team_list from team_list.csv")

    wb.save(workbook_path)
    print(f"saved {workbook_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
